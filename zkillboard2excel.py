import sys
import os
import time
import datetime
import json
import csv
import copy
import urllib.error
import urllib.request
import urllib.parse
import distutils.util
from collections import OrderedDict

import openpyxl

import sde2json


OPTIONS = [
    ('--lang', "The language of a export file. Can use 'de', 'en', 'fr', 'ja', 'ru', 'zh'. default: 'en'"),
    ('--filepath', 'Path to export file. default: export/export'),
    ('--format', "File format to export. Can use 'excel' or 'csv'. default: excel"),
    ('--clear-cache', 'Finally, clear the cache. default: False'),
    ('--update-sde', 'Check for SDE Updates. default: False'),
    ('--page', 'Start page of zKillboard API. default: 1'),
    ('--limit', 'Number of pages read. default: 1'),
]

LANGUAGES = [
    'de',
    'en',
    'fr',
    'ja',
    'ru',
    'zh',
]

TYPES_JSON_PATH = 'types.json'

UNIVERSES_JSON_PATH = 'universes.json'

SETTINGS_JSON_PATH = 'settings.json'
DEFAULT_SETTINGS = {
    'zkb_url': '',
    'lang': 'en',
    'filepath': os.path.join('export', 'export'),
    'format': 'excel',
    'clear-cache': False,
    'update-sde': False,
    'page': 1,
    'limit': 1,
}

CACHED_JSON_PATH = 'cached.json'
CACHED_KEYS = (
    'killmails',
    'characters',
    'corporations',
    'alliances',
)

LOCALE_MENUITEMS_JSON_PATH = 'locale_menuitems.json'
MENUITEMS = (
    'killmail_id',
    'killmail_time',
    'ship',
    'security',
    'region',
    'system',
    'damage',
    'value',
    'points',
    'involved',
    'character',
    'corporation',
    'alliance',
)

FETCH_MODIFIERS = {
    'character': 'characterID',
    'corporation': 'corporationID',
    'alliance': 'allianceID',
    'ship': 'shipTypeID',
    'group': 'groupID',
    'system': 'solarSystemID',
    'constellation': 'constellationID',
    'region': 'regionID',
}

FOCUS_KEYS = [
    'character',
    'corporation',
    'alliance',
    'ship',
    'group',
]

def get_json_by_file(path):
    if os.path.isfile(path):
        with open(path, 'r') as file:
            return json.load(file)
    return {}

SETTINGS = copy.deepcopy(DEFAULT_SETTINGS)
if os.path.isfile(SETTINGS_JSON_PATH):
    for setting_key, setting_value in get_json_by_file(SETTINGS_JSON_PATH).items():
        if isinstance(setting_value, str) and not setting_value:
            continue

        SETTINGS[setting_key] = setting_value

CACHED = get_json_by_file(CACHED_JSON_PATH)
if not CACHED:
    for cached_key in CACHED_KEYS:
        CACHED[cached_key] = {}

TYPES = get_json_by_file(TYPES_JSON_PATH)

UNIVERSES = get_json_by_file(UNIVERSES_JSON_PATH)

LOCALE_MENUITEMS = get_json_by_file(LOCALE_MENUITEMS_JSON_PATH)

def get_json_by_url(url):
    while(True):
        try:
            print('Download: ' + url)
            with urllib.request.urlopen(url) as html:
                data = json.loads(html.read().decode())
            time.sleep(1)
            return data
        except urllib.error.HTTPError:
            print('urllib.error.HTTPError: ' + url)
            time.sleep(30)

def get_link(target_key, target_id, name):
    url = 'https://zkillboard.com/%s/%d/' % (target_key, target_id)
    return '=HYPERLINK("%s", "%s")' % (url, name)

def get_killmail_id(killmail, zkb):
    return killmail['killmail_id']

def get_killmail_id_for_excel(killmail, zkb):
    killmail_id = get_killmail_id(killmail, zkb)
    return get_link('kill', killmail_id, killmail_id)

def get_killmail_time(killmail, zkb):
    date = datetime.datetime.strptime(killmail['killmail_time'], '%Y-%m-%dT%H:%M:%SZ')
    return date.strftime('%Y-%m-%d %H:%M')

def get_ship(killmail, zkb):
    type_id = killmail['ship_id']
    return TYPES['types'][str(type_id)][SETTINGS['lang']]

def get_ship_for_excel(killmail, zkb):
    return get_link('ship', killmail['ship_id'], get_ship(killmail, zkb))

def get_security(killmail, zkb):
    system_id = killmail['system_id']
    return round(UNIVERSES['systems'][str(system_id)]['security'], 1)

def get_region(killmail, zkb):
    return UNIVERSES['regions'][str(killmail['region_id'])][SETTINGS['lang']]

def get_region_for_excel(killmail, zkb):
    return get_link('region', killmail['region_id'], get_region(killmail, zkb))

def get_system(killmail, zkb):
    system_id = killmail['system_id']
    return UNIVERSES['systems'][str(system_id)][SETTINGS['lang']]

def get_system_for_excel(killmail, zkb):
    return get_link('system', killmail['system_id'], get_system(killmail, zkb))

def get_damage(killmail, zkb):
    return killmail['damage_taken']

def get_value(killmail, zkb):
    return zkb['totalValue']

def get_value_for_excel(killmail, zkb):
    return int(get_value(killmail, zkb))

def get_points(killmail, zkb):
    return zkb['points']

def get_involved(killmail, zkb):
    return killmail['involved']

def get_player(player_key, player_id):
    player_id_str = str(player_id)
    esi_url = 'https://esi.evetech.net/latest/%s/%s/' % (player_key, player_id_str)

    if player_id_str in CACHED[player_key]:
        print('Cached: ' + esi_url)
        return CACHED[player_key][player_id_str]

    name = get_json_by_url(esi_url)['name']
    CACHED[player_key][player_id_str] = name
    return name 

def get_character(killmail, zkb):
    name = ''
    if killmail['character_id']:
        name = get_player('characters', killmail['character_id'])
    return name

def get_character_for_excel(killmail, zkb):
    name = get_character(killmail, zkb)
    if name:
        return get_link('character', killmail['character_id'], name)
    return

def get_corporation(killmail, zkb):
    name = ''
    if killmail['corporation_id']:
        name = get_player('corporations', killmail['corporation_id'])
    return name

def get_corporation_for_excel(killmail, zkb):
    name = get_corporation(killmail, zkb)
    if name:
        name = get_link('corporation', killmail['corporation_id'], name)
    return name

def get_alliance(killmail, zkb):
    name = ''
    if killmail['alliance_id']:
        name = get_player('alliances', killmail['alliance_id'])
    return name

def get_alliance_for_excel(killmail, zkb):
    name = get_alliance(killmail, zkb)
    if name:
        name = get_link('alliance', killmail['alliance_id'], name)
    return name

def parse_killmail(killmail):
    ship_id = killmail['victim']['ship_type_id']
    system_id = killmail['solar_system_id']
    constellation_id = UNIVERSES['systems'][str(system_id)]['constellation_id']
    km = {
        'killmail_id': killmail['killmail_id'],
        'killmail_time': killmail['killmail_time'],
        'damage_taken': killmail['victim']['damage_taken'],
        'involved': len(killmail['attackers']),
        'ship_id': ship_id,
        'group_id': TYPES['types'][str(ship_id)]['group_id'],
        'system_id': system_id,
        'constellation_id': constellation_id,
        'region_id':UNIVERSES['constellations'][str(constellation_id)]['region_id'],
    }

    for player_type in ['character', 'corporation', 'alliance']:
        player_key = player_type + '_id'
        if player_key in killmail['victim']:
            km[player_key] = killmail['victim'][player_key]
        else:
            km[player_key] = None

    return km

def get_killmails():
    killmails = OrderedDict()
    header = []
    focus_key = None
    focus_id = None

    for name in MENUITEMS:
        header.append(LOCALE_MENUITEMS[name][SETTINGS['lang']])

    parse_result = urllib.parse.urlparse(SETTINGS['zkb_url'])
    zkb_url = parse_result.scheme + '://' + parse_result.netloc + '/api/'

    for value in parse_result.path.split('/'):
        if not value:
            continue

        if value in FETCH_MODIFIERS:
            zkb_url += FETCH_MODIFIERS[value]
        else:
            zkb_url += value

        zkb_url += '/'

        if not focus_id and focus_key:
            focus_id = int(value)

        if not focus_key and value in FOCUS_KEYS:
            focus_key = value + '_id'

    for limit in range(SETTINGS['limit']):
        url = zkb_url + 'page/%d/' % (limit + SETTINGS['page'])

        for zkb in get_json_by_url(url):
            killmail_id_str = str(zkb['killmail_id'])

            esi_url = 'https://esi.evetech.net/latest/killmails/%d/%s/' % (zkb['killmail_id'], zkb['zkb']['hash'])
            if killmail_id_str in CACHED['killmails']:
                killmail = CACHED['killmails'][killmail_id_str]
            else:
                killmail = parse_killmail(get_json_by_url(esi_url))
                CACHED['killmails'][killmail_id_str] = killmail

            values = []
            for name in MENUITEMS:
                func_name = 'get_' + name

                if SETTINGS['format'] == 'excel' and func_name + '_for_excel' in globals():
                    func_name += '_for_excel'

                values.append(globals()[func_name](killmail, zkb['zkb']))

            killmails[zkb['killmail_id']] = values

        save_cache_json()

    return (killmails, header, focus_key, focus_id)

def zkillboard2csv():
    killmails, header, focus_key, focus_id = get_killmails()

    with open(SETTINGS['filepath'] + '.csv', 'w') as file:
        writer = csv.writer(file, lineterminator='\n')
        writer.writerow(header)
        writer.writerows(killmails.values())

def zkillboard2excel():
    killmails, header, focus_key, focus_id = get_killmails()

    wb = openpyxl.Workbook()
    sheet = wb.active

    for i, header_name in enumerate(header):
        sheet.cell(row=1, column=i+1).value = header_name

    row = 2
    for killmail_id, killmail in killmails.items():
        background = '006400'
        if focus_key and CACHED['killmails'][str(killmail_id)][focus_key] == focus_id:
            background = '8B0000'

        pattern_fill = openpyxl.styles.PatternFill(patternType='solid', fgColor=background, bgColor=background)
        border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='thin', color='000000'),
            bottom=openpyxl.styles.Side(style='thin', color='000000'),
            left=openpyxl.styles.Side(style='thin', color='000000'),
            #right=openpyxl.styles.Side(style='thin', color='000000'),
        )

        for i, value in enumerate(killmail):
            cell = sheet.cell(row=row, column=i+1)
            cell.value = value
            cell.font = openpyxl.styles.fonts.Font(color='FFFFFF')
            cell.fill = pattern_fill
            cell.border = border

        for j in range(3):
            cell = sheet.cell(row=row, column=len(MENUITEMS)+j+1)
            cell.fill = pattern_fill
            cell.border = openpyxl.styles.Border(
                top=openpyxl.styles.Side(style='thin', color='000000'),
                bottom=openpyxl.styles.Side(style='thin', color='000000'),
            )

        row += 1

    wb.save(SETTINGS['filepath'] + '.xlsx')

def save_settings_json():
    with open(SETTINGS_JSON_PATH, 'w', encoding='utf-8') as file:
        json.dump(SETTINGS, file, indent=4)

def save_cache_json():
    with open(CACHED_JSON_PATH, 'w', encoding='utf-8') as file:
        json.dump(CACHED, file)

def clear_cache_json():
    if os.path.isfile(CACHED_JSON_PATH):
        os.remove(CACHED_JSON_PATH)

def run():
    dirname = os.path.dirname(SETTINGS['filepath'])
    if not os.path.isdir(dirname):
        os.makedirs(dirname)

    if SETTINGS['update-sde']:
        sde2json.run()

    if SETTINGS['format'] == 'excel':
        zkillboard2excel()
    else:
        zkillboard2csv()

    if SETTINGS['clear-cache']:
        clear_cache_json()

def command_line():
    if len(sys.argv) < 2:
        print('''
usage: python zkillboard2excel.py zKillboard-URL [options]
Options and arguments:''')

        for option in OPTIONS:
            print('{0:<15}: {1}'.format(option[0], option[1]))
        return

    SETTINGS.update(copy.deepcopy(DEFAULT_SETTINGS))
    SETTINGS['zkb_url'] = sys.argv[1]

    if len(sys.argv) >= 3:
        for argument in sys.argv[2:]:
            try:
                key, value = argument.split('=')

                if key == '--lang':
                    if value in LANGUAGES:
                        SETTINGS['lang'] = value
                    else:
                        print("Does not support '%s' language" % value)
                elif key == '--filepath':
                    SETTINGS['filepath'] = value
                elif key == '--format':
                    if value.lower() in ['excel', 'csv']:
                        SETTINGS['format'] = value.lower()
                    else:
                        print("Does not support '%s' format" % value)
                elif key == '--clear-cache':
                    SETTINGS['clear-cache'] = bool(distutils.util.strtobool(value))
                elif key == '--update-sde':
                    SETTINGS['update-sde'] = bool(distutils.util.strtobool(value))
                elif key == '--page':
                    SETTINGS['page'] = int(value)
                elif key == '--limit':
                    SETTINGS['limit'] = int(value)
                else:
                    print("Option Error: '%s' does not exist" % key)
            except ValueError:
                print('Value Error: %s' % argument)

    save_settings_json()

    run()

if __name__ == '__main__':
    command_line()
